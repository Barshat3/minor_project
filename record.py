from datetime import datetime
from openpyxl import Workbook, load_workbook

def log_to_excel(name):
    file_name = "attendance_log.xlsx"
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        # Load existing workbook or create a new one
        wb = load_workbook(file_name)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        # Create headers if the file doesn't exist
        ws.append(["Name", "Timestamp"])

    # Append the name and timestamp
    ws.append([name, current_time])

    # Save the workbook
    wb.save(file_name)
    print(f"Logged: {name} at {current_time}")